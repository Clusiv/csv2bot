import openpyxl
import shutil
import os

def clear_col():
    wb = openpyxl.load_workbook("app/base.xlsx")
    ws = wb.active

    ws.delete_cols(0)
    ws.delete_cols(2,3)
    ws.delete_cols(3)
    ws.delete_cols(3)

    ws.insert_cols(0)
    count = 1

    for row in range(3, ws.max_row):
        ws.cell(column=1, row=row, value=count)
        count += 1

    return ws

def get_data():
    ws = clear_col()
    return list(ws.iter_rows(values_only=True, min_row=2))
