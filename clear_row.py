import openpyxl
import shutil
import os

def load_wb():
    wb = openpyxl.load_workbook("base.xlsx")
    return wb

def get_ws():
    return load_wb().active

def clear_row():
    wb = load_wb()
    ws = wb.active

    ws.delete_cols(0)
    ws.delete_cols(2,3)
    ws.delete_cols(3)
    ws.delete_cols(3)

    ws.insert_cols(0)
    count = 1
    for row in range(3, ws.max_row):
        ws.cell(column=1, row=row, value=count)
        count += 1

    
    for row in ws.iter_cols(values_only=True, min_col=1, max_col=4, min_row=3, max_row=10):
        print(row)

# clear_row()