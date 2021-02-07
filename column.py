import openpyxl


wb = openpyxl.load_workbook("base.xlsx")
ws = wb.active

# data = list(ws.iter_rows(values_only=True))
# pers = [x for x in data if 'Ибрагим' in x[0]]
pers = []

for col in ws.iter_cols(min_row=1, max_row=1):
    print(col[0].value)

print(pers)